
import re, os, sys, json
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================
# CONFIG: 16 categories + subcategories
# ============================================
CATEGORY_CONFIG = {
    '服装': ['正装','风格','休闲装','运动服','泳装','制服','上衣','外套','腰部','盔甲','裙子','裤子','袜子','材质','花纹','鞋子','靴子','鞋底','领口','手臂服饰','汉服-唐风','汉服-宋风','汉服-明风','汉服-其他','其他服装'],
    '脸部元素': ['面部','眉毛','眼睛','瞳孔','鼻子','嘴巴','牙齿','舌头','脸型'],
    '光照风格': ['光照'],
    '种族生物': ['动物','植物','翅膀','耳朵','精灵/妖精','兽人/亚人','妖怪/怪物','机器人/机甲','天使/恶魔','龙族','不死族/幽灵','幻想生物/其他'],
    '身体': ['皮肤','身材','指甲','肩部','胸部','腰部','腹部','手','其他身体'],
    '视角构图': ['镜头','特写镜头','其他构图','镜头角度','效果','主角动作'],
    'nsfw内容': ['性交','性姿势','射精','私处','裸体','露出','自慰','性玩具','穿孔','束缚','乳房','口交','其他NSFW'],
    '头发头部': ['长度','颜色','发量','发质','整体发型','刘海','侧发','马尾','辫子','发髻','头顶','发饰','发夹','发带','头带','头巾','耳饰','主题发饰','帽子','兜帽','其他装饰','其他头发'],
    '角色tag': ['二次元角色(通用)','崩坏3','绝区零','崩坏星穹铁道|崩铁','原神','碧蓝档案','明日方舟','碧蓝航线','胜利女神妮姬|nikke','鸣潮','魔法少女系列','战双帕弥什','fate系列','偶像大师','车万|东方幻想乡','宝可梦|神奇宝贝','lovelive系列','少女前线','动漫番剧','公主连结|母猪连结','小众手游|小众二游','Hololive','民工漫|火影忍者|海贼王|死神|银魂|七龙珠','舰队collection','赛马娘','虚拟形象','主机游戏|掌机游戏|steam游戏','端游|页游|pc端','精灵角色','兽人/亚人角色','其他角色'],
    '场景建筑': ['室外','城市','室内','地板','家具','床上用品','浴室','季节','天气','大自然','水','天空','云','氛围','学习用品','数码设备','餐具','乐器','武器','食物','其它物品','其他场景'],
    '表情情绪': ['笑','哭','不开心','蔑视','生气','其他表情'],
    '年龄职业': ['人数','年龄','身份'],
    '动作': ['基础动作','手部动作','手部动作(拿着某物)','手部动作(放在某地)','手部动作(抓着某物)','腿部动作','与裙子互动','与裤子互动','与袜子互动','其他动作'],
    '绘图风格': ['画质','艺术风格','艺术类型','艺术派系','艺术家风格','写实','素描','画笔','颜色','魔法特效','其他风格'],
    '配件配饰': ['装饰','首饰','眼镜','面具','手套','头饰','帽子','发饰','小装饰','耳饰','围巾','其他配饰'],
    '背景': ['背景'],
}

CAT_ORDER = list(CATEGORY_CONFIG.keys())

# ============================================
# STEP 1: Parse existing hierarchy
# ============================================
print("[1/6] Parsing existing hierarchy...")
data_dir = r"D:\SillyTavern\SillyTavern\public\scripts\extensions\third-party\chami_tavern-scene-plugin\data"
path_tags_sql = os.path.join(data_dir, 'tags_2025_03_31.sql')
with open(path_tags_sql, 'r', encoding='utf-8') as f:
    content = f.read()

# subgroups: sg_id -> (group_id, sg_name)
subgroups = {}
for m in re.finditer(r'INSERT OR REPLACE INTO "tag_subgroups" .*?VALUES \((.*?)\);', content):
    vals = m.group(1)
    match = re.match(r"(\d+),\s*(\d+),\s*'([^']*)'", vals)
    if match:
        subgroups[int(match.group(1))] = (int(match.group(2)), match.group(3))

# tags: list of (tag_name, translate, sg_id)
hier_tags = []
for m in re.finditer(r'INSERT OR REPLACE INTO "tag_tags" .*?VALUES \((.*?)\);', content):
    vals = m.group(1)
    match = re.match(r"(\d+),\s*(\d+),\s*'([^']*)',\s*'([^']*)'", vals)
    if match:
        sg_id = int(match.group(2))
        hier_tags.append((match.group(3), match.group(4), sg_id))

print(f"  Loaded {len(hier_tags)} hierarchical tags, {len(subgroups)} subgroups")

# ============================================
# STEP 2: Parse danbooru tags
# ============================================
print("[2/6] Parsing danbooru tags...")
danbooru_files = sorted([f for f in os.listdir(data_dir) if f.startswith('danbooru_') and f.endswith('.sql')])
danbooru_tags = {}  # tag_name -> (translate, hot)

for i, fname in enumerate(danbooru_files):
    fpath = os.path.join(data_dir, fname)
    with open(fpath, 'r', encoding='utf-8') as f:
        fc = f.read()
    count = 0
    for m in re.finditer(r"INSERT OR REPLACE INTO\s+\"danbooru_tag\".*?VALUES\s*\((.*?)\);", fc):
        vals = m.group(1)
        match = re.match(r"(\d+),\s*'([^']*)',\s*(\d+),\s*'([^']*)',\s*(\d+)", vals)
        if match:
            tn = match.group(2)
            trans = match.group(4)
            hot = int(match.group(5))
            if tn not in danbooru_tags:
                danbooru_tags[tn] = (trans, hot)
            count += 1
    if (i+1) % 5 == 0:
        print(f"  ... {i+1}/{len(danbooru_files)} files")

print(f"  Loaded {len(danbooru_tags)} unique danbooru tags")

# ============================================
# STEP 3: Build classification rules
# ============================================

# --- 3a: Existing hierarchy → new category mapping ---
def map_hier_to_new(sg_id, tag_name):
    """Map existing subgroup_id to (cat1, cat2) based on final config"""
    if sg_id not in subgroups:
        return None, None
    gid, sg_name = subgroups[sg_id]

    # Group 1: 人物
    if gid == 1:
        m = {
            1: ('年龄职业', '人数'),       # 对象 (1girl/2girls/solo等)
            2: ('年龄职业', '身份'),       # 身份
            3: ('角色tag', '二次元角色(通用)'), # 二次元角色
            4: ('年龄职业', '年龄'),       # 年龄
            5: ('身体', '皮肤'),
            6: ('身体', '身材'),
            7: ('脸部元素', '脸型'),
            8: ('头发头部', None),          # 头发→待后续拆分
            9: ('脸部元素', '面部'),
            10: ('种族生物', '耳朵'),
            11: ('脸部元素', '眉毛'),
            12: ('脸部元素', '眼睛'),
            13: ('脸部元素', '瞳孔'),
            14: ('脸部元素', '鼻子'),
            15: ('脸部元素', '嘴巴'),
            16: ('脸部元素', '牙齿'),
            17: ('脸部元素', '舌头'),
            18: ('身体', '指甲'),
            19: ('身体', '肩部'),
            20: ('身体', '胸部'),
            21: ('身体', '腰部'),
            22: ('身体', '腹部'),
            23: ('种族生物', '翅膀'),
        }
        return m.get(sg_id, (None, sg_name))

    # Group 2: 服饰
    if gid == 2:
        # Action-related subgroups go to 动作
        if sg_id in [36]: return ('动作', '与裙子互动')
        if sg_id in [38]: return ('动作', '与裤子互动')
        if sg_id in [40]: return ('动作', '与袜子互动')
        # Accessories
        if sg_id in [42]: return ('配件配饰', '装饰')
        if sg_id in [48]: return ('配件配饰', '首饰')
        if sg_id in [49]: return ('配件配饰', '眼镜')
        if sg_id in [50]: return ('配件配饰', '面具')
        if sg_id in [53]: return ('配件配饰', '手套')
        if sg_id in [54]: return ('配件配饰', '头饰')
        if sg_id in [55]: return ('配件配饰', '帽子')
        if sg_id in [56]: return ('配件配饰', '发饰')
        if sg_id in [57]: return ('配件配饰', '小装饰')
        if sg_id in [58]: return ('配件配饰', '耳饰')
        if sg_id in [59]: return ('配件配饰', '围巾')
        # Arms/hands
        if sg_id in [51]: return ('服装', '手臂服饰')
        if sg_id in [52]: return ('身体', '手')
        # Everything else is 服装
        return ('服装', sg_name)

    # Group 3: 表情动作
    if gid == 3:
        if sg_id in [60,61,62,63,64,65]: return ('表情情绪', sg_name)
        return ('动作', sg_name)

    # Group 4: 画面
    if gid == 4:
        if sg_id == 78: return ('光照风格', '光照')
        if sg_id == 83: return ('背景', '背景')
        return ('绘图风格', sg_name)

    # Group 5: 环境 → 场景建筑
    if gid == 5: return ('场景建筑', sg_name)

    # Group 6: 场景 → 场景建筑
    if gid == 6: return ('场景建筑', sg_name)

    # Group 7: 物品
    if gid == 7:
        if sg_id in [105]: return ('种族生物', '动物')
        if sg_id in [106]: return ('种族生物', '植物')
        return ('场景建筑', sg_name)

    # Group 8: 镜头 → 视角构图
    if gid == 8: return ('视角构图', sg_name)

    # Group 9: 汉服 → 服装
    if gid == 9:
        sname = sg_name.rstrip(':')
        if '唐' in sname: return ('服装', '汉服-唐风')
        if '宋' in sname: return ('服装', '汉服-宋风')
        if '明' in sname: return ('服装', '汉服-明风')
        return ('服装', '汉服-其他')

    # Group 10: 魔法系 → 绘图风格.魔法特效
    if gid == 10: return ('绘图风格', '魔法特效')

    return (None, sg_name)

# --- 3b: Hair tag splitting rules ---
HAIR_RULES = [
    # (keyword_pattern, subcategory) - checked in order
    (r'^(long_hair|short_hair|medium_hair|very_long_hair)$', '长度'),
    (r'^(blonde_hair|brown_hair|black_hair|white_hair|red_hair|blue_hair|pink_hair|purple_hair|green_hair|silver_hair|grey_hair|gradient_hair|multicolored_hair|two_tone_hair)$', '颜色'),
    (r'.*(hair_between_eyes|hair_over_eyes|hair_over_one_eye|swept_bangs|parted_bangs|blunt_bangs|asymmetrical_bangs|wispy_bangs|curtained_hair|sidelocks).*', '刘海'),
    (r'.*(ponytail|side_ponytail|high_ponytail|low_ponytail|short_ponytail|french_braid|twin_braids|single_braid|braid|side_braid|crown_braid).*', '马尾' if 'ponytail' in r'ponytail' else '辫子'),
    (r'.*(twintails|low_twintails|high_twintails|short_twintails|triple_tails|quad_tails).*', '整体发型'),
    (r'.*(bun|double_bun|triple_bun|heart_bun|odango).*', '发髻'),
    (r'.*(ahoge|antenna_hair|hair_ribbon|hair_bow|hairband|hair_ornament|hair_flower|hair_ring|hair_bell|hair_feather).*', '发饰'),
    (r'.*(hair_clip|hairpin|bobby_pin).*', '发夹'),
    (r'.*(headband|head_ribbon).*', '头带'),
    (r'.*(bandana|kerchief).*', '头巾'),
    (r'.*(hood|hood_up|hood_down).*', '兜帽'),
    (r'.*(hat|cap|beret|fedora|sun_hat|witch_hat|top_hat|straw_hat|bucket_hat|cowboy_hat|newsboy_cap|peaked_cap|baseball_cap|santa_hat|party_hat|animal_hat|bear_hat|cat_hat|bunny_hat).*', '帽子'),
    (r'.*(earrings|ear_ring|ear_piercing).*', '耳饰'),
    (r'.*(drill_hair|curly_hair|wavy_hair|spiky_hair|messy_hair|straight_hair|fluffy_hair).*', '发质'),
    (r'.*(thick_hair|thin_hair|bald|receding_hairline).*', '发量'),
    (r'.*(hair_bun|hair_up|hair_down|hair_pulled_back|hair_over_shoulder|hair_between_breasts|floating_hair).*', '整体发型'),
    (r'.*(side_hair|asymmetrical_hair).*', '侧发'),
    (r'.*(animal_hair|hair_theme|colored_hair_tips|streaked_hair|dyed_hair).*', '主题发饰'),
    (r'.*(hair|wig).*', '其他头发'),  # fallback for any hair tag
]

# Hack fix for ponytail rule
HAIR_RULES_FIXED = [
    (['long_hair','short_hair','medium_hair','very_long_hair'], '长度'),
    (['blonde_hair','brown_hair','black_hair','white_hair','red_hair','blue_hair','pink_hair','purple_hair','green_hair','silver_hair','grey_hair','gray_hair','gradient_hair','multicolored_hair','two_tone_hair','colored_hair','orange_hair','teal_hair','aqua_hair','light_blonde_hair','dark_hair','light_hair'], '颜色'),
    (['hair_between_eyes','hair_over_eyes','hair_over_one_eye','swept_bangs','parted_bangs','blunt_bangs','asymmetrical_bangs','wispy_bangs','curtained_hair','sidelocks','front_hair'], '刘海'),
    (['ponytail','side_ponytail','high_ponytail','low_ponytail','short_ponytail','long_ponytail','front_ponytail','split_ponytail','folded_ponytail'], '马尾'),
    (['braid','braided_hair','french_braid','twin_braids','single_braid','side_braid','crown_braid','dutch_braid','fishtail_braid','multiple_braids','long_braid'], '辫子'),
    (['twintails','low_twintails','high_twintails','short_twintails','triple_tails','quad_tails','double_bun','triple_bun','half_up_hair','half_updo'], '整体发型'),
    (['bun','heart_bun','odango','side_bun','messy_bun','ballerina_bun','hair_bun','updo','chignon'], '发髻'),
    (['ahoge','antenna_hair','hair_ribbon','hair_bow','hairband','hair_ornament','hair_flower','hair_ring','hair_bell','hair_feather','hair_scrunchy'], '发饰'),
    (['hair_clip','hairpin','bobby_pin','barrette','hair_tie','hair_stick'], '发夹'),
    (['headband','head_ribbon','hair_band','head_dress','headpiece','tiara','crown','circlet'], '头带'),
    (['bandana','kerchief','bandanna','head_scarf'], '头巾'),
    (['hood','hood_up','hood_down','hoodie'], '兜帽'),
    (['hat','cap','beret','fedora','sun_hat','witch_hat','top_hat','straw_hat','bucket_hat','cowboy_hat','newsboy_cap','peaked_cap','baseball_cap','santa_hat','party_hat','animal_hat','bear_hat','cat_hat','bunny_hat','visor','beanie','bowler_hat','sombrero','pirate_hat','chef_hat','helmet','bike_helmet'], '帽子'),
    (['earrings','ear_ring','ear_piercing','ear_cuffs'], '耳饰'),
    (['drill_hair','curly_hair','wavy_hair','spiky_hair','messy_hair','straight_hair','fluffy_hair','frizzy_hair','glossy_hair','shiny_hair','wet_hair'], '发质'),
    (['thick_hair','thin_hair','bald','receding_hairline','thick_eyebrows','thin_eyebrows'], '发量'),
    (['hair_over_shoulder','hair_between_breasts','floating_hair','hair_pulled_back','hair_down','hair_up','hair_back'], '整体发型'),
    (['side_hair','asymmetrical_hair','side_locks'], '侧发'),
    (['animal_hair','colored_hair_tips','streaked_hair','dyed_hair','inner_colored_hair','two_tone_hair','multicolored_hair','gradient_hair','hair_highlights','colored_inner_hair'], '主题发饰'),
    (['maid_headdress','nurse_cap','military_hat','police_hat','sailor_hat','head_veil','bridal_veil','veil'], '其他装饰'),
    (['hair'], '其他头发'),
]

def classify_hair_tag(tag_name):
    """Classify a hair-related tag into subcategory"""
    tag_low = tag_name.lower()
    for keywords, subcat in HAIR_RULES_FIXED:
        for kw in keywords:
            if kw == tag_low or (len(kw) > 4 and kw in tag_low):
                return subcat
    return '其他头发'

# --- 3c: NSFW detection rules ---
NSFW_KEYWORDS = {
    '性交': ['penetration','intercrural','paizuri','footjob','handjob','fingering','threesome','orgy','double_penetration','anal','vaginal','sex','intercrural_sex','thigh_sex','tail_sex','armpit_sex','womb_tattoo','after_sex','after_anal','after_vaginal','creampie','internal_cum','cum_in_pussy','cum_in_ass','cum_in_mouth','cum_in_womb','nakadashi','defloration','first_sex','deflowering','fellatio','cunnilingus','irrumatio','deep_throat','blowjob','blowjob_face','fellatio_face'],
    '性姿势': ['spread_legs','legs_up','legs_over_head','mating_press','full_nelson','doggystyle','cowgirl_position','reverse_cowgirl','missionary','suspended_congress','pile_driver','stand_and_carry','spoon_position','prone_bone','amazon_position','back_to_back','standing_sex','sex_from_behind','all_fours','on_bed','on_table','on_desk','on_chair','against_wall','against_window','on_floor','in_bath','in_pool','in_water','on_grass','outdoor_sex','public_sex','sex_machine','suspended','inverted','upside_down'],
    '射精': ['cum','semen','ejaculation','precum','cumshot','facial','cum_on_body','cum_on_face','cum_on_hair','cum_on_clothes','cum_on_breasts','cum_on_stomach','cum_on_back','cum_on_feet','cum_on_hands','cum_in_uterus','cum_pool','cum_drip','cum_drool','cum_string','cum_bubble','excessive_cum','huge_cum','cum_leaking','cum_inflation','cum_bath','cum_covered','cum_stain','cum_on_food','cum_in_food','cum_in_drink','cum_on_others','cum_swap','snowballing','cum_kiss','gokkun','bukkake','multiple_cumshots','messy_facial','ruined_facial'],
    '私处': ['pussy','vagina','clitoris','clit','labia','cameltoe','cunt','penis','cock','dick','balls','testicles','scrotum','glans','foreskin','shaft','penis_tip','urethra','perineum','taint','balls_touch','balls_resting','balls_deep','uncircumcised','circumcised','erection','erect_penis','flaccid','semi_erect','big_penis','small_penis','huge_penis','monster_cock','horsecock','alien_penis','tentacle_penis','dildo_penis','clothed_penis','pantyshot','upskirt','downblouse','pussy_juice','pussy_juice_drip','pussy_juice_string','pussy_juice_pool','groin','crotch'],
    '裸体': ['nude','naked','topless','bottomless','completely_nude','shirtless','pantless','no_panties','no_bra','see_through','sheer_clothes','sheer_underwear','sheer_shirt','sheer_dress','sheer_skirt','translucent_clothing','clothes_lift','shirt_lift','skirt_lift','dress_lift','dress_pull','skirt_pull','shirt_pull','undressing','stripping','clothing_removal','clothes_removed','wardrobe_malfunction','clothes_rip','clothes_torn','exposed_breasts','exposed_nipples','exposed_pussy','exposed_penis','exposed'],
    '露出': ['exhibitionism','public_nudity','flashing','streaking','caught','embarrassed_nude','embarrassed','caught_masturbating','caught_sex','outdoor_nude','beach_nude','park_nude','forest_nude','street_nude','changing_room','locker_room','bathhouse','onsen','mixed_bathing','skinny_dipping','nude_beach','nudist','exposed_breasts_outdoors','exposed_outdoors','risky','risky_outdoors','window','balcony_nude'],
    '自慰': ['masturbation','masturbating','fingering_self','dildo_masturbation','vibrator_masturbation','pillow_humping','masturbating_in_bath','masturbating_in_bed','hands-free_masturbation','orgasm','climax','squirting','female_ejaculation','male_masturbation','female_masturbation','multiple_orgasm','forced_orgasm','edge','edging','orgasm_denial','ruined_orgasm','mind_break_orgasm','ahegao','rolling_eyes','crossed_eyes','tongue_out','drooling'],
    '性玩具': ['dildo','vibrator','sex_toy','anal_beads','butt_plug','tail_plug','vibrating_plug','remote_vibrator','wand_vibrator','hitachi','rabbit_vibrator','bullet_vibrator','clit_vibrator','fleshlight','onahole','cock_ring','ball_gag','gag','ring_gag','tape_gag','cleave_gag','otm_gag','mouth_plug','bit_gag','spreader_bar','collar','leash','whip','crop','paddle','flogger','cane','riding_crop','dildo_harness','strap_on','double_dildo','dildo_in_pussy','dildo_in_ass','vibrator_in_pussy','vibrator_in_ass','dildo_sucking'],
    '穿孔': ['piercing','pierced','nipple_piercing','clit_piercing','tongue_piercing','nose_piercing','belly_piercing','ear_piercing','lip_piercing','septum_piercing','navel_piercing','genital_piercing','prince_albert','body_piercing','earring'],
    '束缚': ['bound','tied','tied_up','shibari','bondage','restrained','handcuffs','cuffed','shackles','chains','rope','bound_wrists','bound_legs','bound_arms','bound_feet','hogtie','spreadeagle','suspended_bondage','straitjacket','tape_bondage','chair_bondage','pillory','stocks','cage','jail','prison','cell','dungeon','captive','prisoner','kidnapped','hostage','gagged','blindfold','blindfolded','leash_and_collar','chained','shackled','cuffed_hands','cuffed_legs','handcuffs_behind_back','arms_behind_back','wrists_tied','ankles_tied','knees_tied'],
    '乳房': ['breasts','nipples','cleavage','areola','areolae','breast','nipple','bust','boobs','boob','tit','tits','bosom','underboob','sideboob','nipple_slip','pokies','perky_nipples','erect_nipples','puffy_nipples','inverted_nipples','large_nipples','small_nipples','dark_nipples','pink_nipples','nipple_pull','nipple_tweak','nipple_pinch','nipple_squeeze','nipple_play','breast_squeeze','breast_hold','breast_grab','breast_smother','breast_press','breast_sucking','breast_feeding','lactation','milking','breast_milk','milk','huge_breasts','large_breasts','small_breasts','flat_chest','breast_size','breasts_together','breasts_apart','breasts_squeezed','breasts_resting','mammaries'],
    '口交': ['fellatio','blowjob','cunnilingus','irrumatio','deep_throat','deepthroat','fellatio_face','blowjob_face','cum_in_mouth','licking_penis','licking_pussy','licking_nipple','licking_balls','rimming','anilingus','face_fuck','throat_bulge','throat_fuck'],
    '其他NSFW': ['nsfw','hentai','ecchi','erotic','porn','pornography','sexual','sensual','lewd','lewd_gesture','lewd_position','lewd_face','provocative','seductive','suggestive'],
}

def classify_nsfw(tag_name):
    for subcat, keywords in NSFW_KEYWORDS.items():
        for kw in keywords:
            if kw == tag_name.lower() or (len(kw) > 4 and kw in tag_name.lower()):
                return subcat
    return None

# --- 3d: IP/copyright detection ---
IP_MAP = {
    '崩坏3': ['honkai_impact','honkai_impact_3rd','honkai_impact_3','hi3','kiana_kaslana','raiden_mei','bronya_zaychik','himeko_murata','seele_volerei','fu_hua','theresa_apocalypse','rita_rossweisse','durandal','elysia','mobius','aponia','eden','pardo','griseo','kosma','vill_v','kevin_kaslana','su_(honkai)','kalpas','sakura_(honkai)','honkai_gakuen','ggz'],
    '绝区零': ['zenless_zone_zero','zzz','belle_(zzz)','wise_(zzz)','nicole_demara','anby_demara','billy_kid','nekomiya_mana','corin_wickes','ellen_joe','von_lycaon','grace_howard','anton_ivanov','ben_bigger','koleda_belobog','soldier_11_(zzz)','zhu_yuan','qingyi','miyabi_(zzz)','soukaku','yanagi_(zzz)','harumasa_(zzz)','lighter_(zzz)','caesar_king','burnice_white','piper_wheel'],
    '崩坏星穹铁道|崩铁': ['honkai_star_rail','hsr','march_7th','dan_heng','himeko_(star_rail)','welt_(star_rail)','stelle_(star_rail)','caelus_(star_rail)','seele_(star_rail)','bronya_(star_rail)','jing_yuan','kafka_(star_rail)','blade_(star_rail)','silver_wolf','luocha','fu_xuan','jingliu','topaz','huohuo','ruan_mei','dr_ratio','sparkle','black_swan','acheron','aventurine','robin','boothill','firefly','jade_(star_rail)','feixiao','lingsha','rappa','sunday_(star_rail)','fugue','the_herta','aglaea','tribbie','mydei','anaxa','castorice','phainon','cyrene'],
    '原神': ['genshin_impact','paimon_(genshin)','aether_(genshin)','lumine_(genshin)','amber_(genshin)','kaeya','lisa_(genshin)','barbara_(genshin)','razor_(genshin)','xiangling','beidou','xingqiu','ningguang','fischl_(genshin)','bennett_(genshin)','noelle_(genshin)','chongyun','sucrose_(genshin)','jean_(genshin)','diluc','qiqi','mona_(genshin)','keqing','venti','klee','diona','tartaglia','zhongli','xinyan','albedo','ganyu','xiao_(genshin)','hu_tao','rosaria','yanfei','eula','kaedehara_kazuha','ayaka','yoimiya','sayu','raiden_shogun','kujou_sara','kokomi','thoma','gorou','itto','shenhe','yunjin','yae_miko','ayato','yelan','kuki_shinobu','heizou','collei','tighnari','dori','nilou','cyno','candace','nahida','layla','faruzan','wanderer','alhaitham','yaoyao','dehya','mika','baizhu','kaveh','kirara','lyney','lynette','freminet','neuvillette','wriothesley','furina','charlotte_(genshin)','navia','chevreuse','xianyun','gaming_(genshin)','chiori_(genshin)','arlechinno','sethos','clorinde','sigewinne','emilie','mualani','kachina','kinich','xilonen','citlali','ororon','chasca_(genshin)','mavuika','lanyan','mizuki_(genshin)','iansan','varesa','genshin','sumeru','inazuma','liyue','mondstadt','fontaine','natlan','snezhnaya','fatui','harbinger','archon','adeptus','yaksha','vision_(genshin)'],
    '碧蓝档案': ['blue_archive','arisu_(blue_archive)','shiroko','hoshino','hina_(blue_archive)','ako_(blue_archive)','iroha_(blue_archive)','mika_(blue_archive)','saori_(blue_archive)','atsuko_(blue_archive)','misaki_(blue_archive)','hiyori_(blue_archive)','azusa_(blue_archive)','mashiro_(blue_archive)','tsurugi','hasumi','nonomi','serika','ayane_(blue_archive)','yuuka','kotama','hinata_(blue_archive)','hanako_(blue_archive)','mari_(blue_archive)','ui_(blue_archive)','sena_(blue_archive)','koharu_(blue_archive)','haruna_(blue_archive)','fuuka_(blue_archive)','junko_(blue_archive)','izumi_(blue_archive)','akari_(blue_archive)','karin_(blue_archive)','asuna_(blue_archive)','neru','tok_(blue_archive)','midori_(blue_archive)','momoi','yuzu_(blue_archive)','reisa','wakamo','mutsuki_(blue_archive)','arona','plana','phrenapates','sensei_(blue_archive)'],
    '明日方舟': ['arknights','amiya_(arknights)','texas_(arknights)','exusiai','lappland','skadi_(arknights)','chen_(arknights)','hoshiguma','mostima','w_(arknights)','surtr','mudrock_(arknights)','kal\'tsit','saria','silence_(arknights)','ifrit','eyjafjalla','siege_(arknights)','blaze_(arknights)','specter_(arknights)','gladiia','skadi_the_corrupting_heart','nearl_(arknights)','blemishine','whislash','margaret_nearl','rosmontis','goldenglow','pozyomka','gavial_(arknights)','gavial_the_invincible','reed_(arknights)','reed_the_flame_shadow','lin_(arknights)','penance_(arknights)','executor_(arknights)','executor_the_ex_foedere','virtuosa','ho_olheyak','typhon_(arknights)','muelsyse','swire_(arknights)','swire_the_elegant_wit','eyjafjalla_the_hvit_aska','shu_(arknights)','ray_(arknights)','ela_(arknights)','ash_(arknights)','frost_(arknights)','tachanka_(arknights)','blitz_(arknights)','fiammetta','horn_(arknights)','irene_(arknights)','saileach','lumen_(arknights)','gnosis','ling_(arknights)','nymph','logos_(arknights)','wisadel','dusk_(arknights)','chongyue','nian_(arknights)'],
    '碧蓝航线': ['azur_lane','enterprise_(azur_lane)','belfast_(azur_lane)','atago_(azur_lane)','takao_(azur_lane)','prinz_eugen_(azur_lane)','illustrious_(azur_lane)','unicorn_(azur_lane)','akagi_(azur_lane)','kaga_(azur_lane)','amagi_(azur_lane)','shinano_(azur_lane)','new_jersey_(azur_lane)','friedrich_der_grosse','agir','hakuryuu','august_von_parseval','ulrich_von_hutten','plymouth_(azur_lane)','hindenburg_(azur_lane)','guam_(azur_lane)','laffey_(azur_lane)','ayanami_(azur_lane)','javelin_(azur_lane)','z23_(azur_lane)','nimi_(azur_lane)','helena_(azur_lane)','cleveland_(azur_lane)','saint_louis_(azur_lane)','baltimore_(azur_lane)','bremerton'],
    '胜利女神妮姬|nikke': ['nikke','goddess_of_victory_nikke','marian_(nikke)','raputre_(nikke)','anis_(nikke)','neon_(nikke)','rupee_(nikke)','privaty','admi','alice_(nikke)','liter','snow_white_(nikke)','scarlet_(nikke)','rapunzel_(nikke)','dorothy_(nikke)','red_hood_(nikke)','modernia','nihilister','d_(nikke)','volume_(nikke)','centi_(nikke)','cocoa_(nikke)','soda_(nikke)','noir_(nikke)','blanc_(nikke)','makima_(nikke)','power_(nikke)','2b_(nikke)','a2_(nikke)','helm_(nikke)','mast_(nikke)','anchor_(nikke)','sakura_(nikke)','moran_(nikke)'],
    '鸣潮': ['wuthering_waves','rover_(wuwa)','yangyang','chixia','baizhi_(wuwa)','yuanwu','taoqi','aalto_(wuwa)','encore_(wuwa)','sanhua','danjin','jianxin','jiyan','yinlin','verina','lingyang','mortefi','changli','jinhsi','zhezhi','xiangli_yao','shorekeeper','youhu','camellya','carlotta','roccia_(wuwa)','brant','phoebe_(wuwa)','zani_(wuwa)','wuwa'],
    '魔法少女系列': ['mahou_shoujo','magical_girl','madoka_magica','puella_magi','kaname_madoka','akemi_homura','miki_sayaka','tomoe_mami','sakura_kyouko','nagisa_momoe','nanoha','fate_testarossa','hayate_yagami','magical_girl_lyrical_nanoha','cardcaptor_sakura','kinomoto_sakura','sakura_card_captor','precure','pretty_cure','cure','magical_girl_raising_project','yuki_yuna','prisma_illya','magical_girl_site','magical_girl_spec_ops_asuka','kill_la_kill','senketsu','ryuko_matoi','symphogear'],
    '战双帕弥什': ['punishing_gray_raven','pgr','lucia_(pgr)','liv_(pgr)','lee_(pgr)','nanami_(pgr)','bianca_(pgr)','alpha_(pgr)','vera_(pgr)','selena_(pgr)','luna_(pgr)','qu_(pgr)','watanabe_(pgr)','kamui_(pgr)','chrome_(pgr)','roland_(pgr)','lamia_(pgr)','no_21_(pgr)','noan_(pgr)','pulao_(pgr)','alisa_(pgr)','bambinata_(pgr)','karenina_(pgr)','ros_(pgr)','teddy_(pgr)','bridget_(pgr)','wanshi_(pgr)','hanying_(pgr)','lilith_(pgr)'],
    'fate系列': ['fate','saber_(fate)','arthur_pendragon','ar','em','shirou_em','rin_tohs','sakura_matou','il','gilgamesh_(fate)','cu_chu','medusa_(fate)','medea_(fate)','heracles_(fate)','iskandar','waver_velvet','jeanne_dal','mordred_(fate)','astolfo_(fate)','nero_clau','tamamo','mash_kyr','scathach_(fate)','bb_(fate)','meltry','passionlip','ereshkigal_(fate)','ishtar_(fate)','musashi_mi','okita_sou','morgan_(fate)','artoria_caster','oberon_(fate)','melus','barghest_(fate)','baobhan_sith','koyans','castoria','fate_grand_order','fgo','fate_stay_night','fate_zero','fate_apocrypha','fate_extra','fate_hollow_atar','fate_strange_fake','tsukihime','ar','shiki_ry','arcueid','ciel_(tsukihime)','akiha_tohn','melty_blood','kohaaku','hisui_(tsukihime)','mahoutsukai_no_yoru','aoko_aoz','alice_k','soujuur','witch_on_the_holy_night'],
    '偶像大师': ['idolmaster','im@s','cg','shiny_colors','million_live','sidem','amami_haru','kisaragi_chi','takatsuki_ya','shijou_taka','shimamura_uz','honda_mio','s','ogata_ch','moroboshi_ki','koshimizu_sa','ichinose_shi','miyamoto_fre','mika_jouga','anastasia_(im@s)','rin_shibu','uzuki_shim','mio_honda','kanade_hay','kaede_taka','rank_(im@s)','asuka_n','fumika_sa','syuko_sh','sae_koba','yoshino_yo','yuki_hime','miho_koh','chihaya_ki','azusa_mi','miki_hos','iroha_(im@s)','amana_osa','tenka_osa','chiyoko_so','asahi_se','fuyuko_ma','mamimi_tan','sakuya_sh','nichika_na','mikoto_ak','luca_ika','nana_abe'],
    '车万|东方幻想乡': ['touhou','hakurei_rei','kirisame_ma','izayoi_sa','remilia_sca','flandre_sca','patchouli_know','sakuya_iz','hong_mei','cirno_(touhou)','alice_marg','yakumo_yu','yakumo_ra','reisen_ud','komeiji_sa','komeiji_k','fujiwara_no_m','inaba_tewu','kochiya_sa','moriya_su','yasaka_kan','hijiri_bya','toyosatomimi_no_m','kaku_se','hata_no_k','houraisan_k','inaba_mokuho'],
    '宝可梦|神奇宝贝': ['pokemon','pikachu','eevee','charizard','mewtwo','gardevoir','lucario','serena_(pokemon)','misty_(pokemon)','may_(pokemon)','dawn_(pokemon)','rosa_(pokemon)','hikari_(pokemon)','nemona','iono','marnie_(pokemon)','cynthia_(pokemon)','lusamine','lillie_(pokemon)','gloria_(pokemon)','leaf_(pokemon)','hilda_(pokemon)','poke'],
    'lovelive系列': ['love_live','school_idol_festival','kousaka_ho','sonoda_u','minami_ko','nishikino_ma','hoshizora_ri','koizumi_ha','yazawa_ni','toujou_n','ayase_el','tsushima_y','watanabe_y','sakurauchi_r','matsuura_k','kurosawa_d','kurosawa_r','kunikida_h','ohara_m','takami_ch','kanan_ma','dia_kuro','you_watan','riko_sak','mari_oha','kanata_kon','ayumu_ue','setsuna_yu','shioriko_mi','ren_haz','chisato_ar','sumire_he','kinako_sa','lanzhu_zh','margarethe_w','natsumi_o','keke_ta','mei_yo','shiki_wa','yu_taki','ai_miy','shizuku_os','karin_as','emma_ve','rina_ten','mia_ta'],
    '少女前线': ['girls_frontline','m4a1_(girls_frontline)','m16a1_(girls_frontline)','sopmod_ii','ar_15_(girls_frontline)','ro635','ump45','ump9','hk416_(girls_frontline)','g11_(girls_frontline)','wa2000_(girls_frontline)','springfield_(girls_frontline)','kar98k_(girls_frontline)','ntw_20_(girls_frontline)','m200_(girls_frontline)','g36_(girls_frontline)','aug_(girls_frontline)','type_95_(girls_frontline)','type_97_(girls_frontline)','ots_14_(girls_frontline)','as_val_(girls_frontline)','vector_(girls_frontline)','thompson_(girls_frontline)','ppsh_41_(girls_frontline)','sten_mkii','micro_uzi_(girls_frontline)','thunder_(girls_frontline)','spas_12_(girls_frontline)','m870_(girls_frontline)','sat8_(girls_frontline)','uzi_(girls_frontline)','fn_fal_(girls_frontline)','hanyang_88_(girls_frontline)','type_56_(girls_frontline)','girls_frontline_2','exilium','groza_(gfl2)','nemesis_(gfl2)','colphne_(gfl2)','cheeta_(gfl2)','krolik_(gfl2)','lotta_(gfl2)','suomi_(gfl2)','ullrid_(gfl2)'],
    '公主连结|母猪连结': ['princess_connect','princess_connect_re_dive','pecorine','karyl','kyaru','kokkoro','yuuki_(princess_connect)','makoto_(princess_connect)','shizuru_(princess_connect)','rino_(princess_connect)','yui_(princess_connect)','hiyori_(princess_connect)','rei_(princess_connect)','shinobu_(princess_connect)','akino_(princess_connect)','mifuyu_(princess_connect)','mahiru_(princess_connect)','yukari_(princess_connect)','saren_(princess_connect)','nozomi_(princess_connect)','chika_(princess_connect)','tsumugi_(princess_connect)','ayane_(princess_connect)','kurumi_(princess_connect)','kasumi_(princess_connect)','yori_(princess_connect)_akai','aoi_(princess_connect)','hatsune_(princess_connect)','shiori_(princess_connect)','io_(princess_connect)','suzuna_(princess_connect)','misato_(princess_connect)','ninon_(princess_connect)','eriko_(princess_connect)','suzume_(princess_connect)','kaori_(princess_connect)','lima_(princess_connect)','tamaki_(princess_connect)','miyako_(princess_connect)','maho_(princess_connect)','anna_(princess_connect)','nanaka_(princess_connect)','monika_(princess_connect)','ayumi_(princess_connect)','rima_(princess_connect)','kyoka_(princess_connect)','illya_(princess_connect)','jun_(princess_connect)','christina_(princess_connect)','muimi_(princess_connect)','shefi_(princess_connect)','ame_(princess_connect)','labyrista','neneka_(princess_connect)','cred_(princess_connect)','ranpha_(princess_connect)','priconne'],
    'Hololive': ['hololive','tokino_sora','roboco_san','sakura_miko','hoshimachi_suisei','azki','shirakami_fubuki','natsuiro_matsuri','akai_haato','haachama','minato_aqua','murasaki_shion','nakiri_ayame','yuzuki_choco','oozora_subaru','ookami_mio','nekomata_okayu','inugami_korone','shiranui_flare','shirogane_noel','houshou_marine','usada_pekora','uruha_rushia','shishiro_botan','ogayu_marin','momosuzu_nene','yukihana_lamy','holy_live','gawr_gura','watson_amelia','takanashi_kiara','mori_calliope','n','irys','tsukumo_sana','ceres_fauna','ouro_kroni','nanashi_mu','hakos_baelz','takane_lui','hakui_ko','sakamata_chloe','kazama_iroha','laplus_darkness','holostars','yagoo','shinove','a_chan_(hololive)','nodoka_(hololive)','harusaki_nodoka'],
    '民工漫|火影忍者|海贼王|死神|银魂|七龙珠': ['naruto','uzumaki_naruto','uchiha_sasuke','haruno_sakura','hatake_kakashi','hyuuga_hinata','nara_shikamaru','akimichi_chouji','yamanaka_ino','rock_lee','tenten','neji_hyuuga','gaara','tsunade_(naruto)','jiraiya','orochimaru_(naruto)','itachi_uchiha','pain_(naruto)','minato_namikaze','kushina_uzumaki','boruto','sarada_uchiha','mitsuki_(boruto)','one_piece','monkey_d._luffy','roronoa_zoro','nami_(one_piece)','usopp','sanji_(one_piece)','tony_tony_chopper','nico_robin','franky_(one_piece)','brook_(one_piece)','j','portgas_d._ace','boa_hanc','trafalgar_law','donquixote_doflamingo','shanks_(one_piece)','buggy_(one_p.","nefertari_vivi","yamato_(one_piece)','carrot_(one_piece)','bleach','kurosaki_ich','kuchiki_ru','ino','ishida_ur','yasutora_sa','abara_ren','hitsugaya_toush','zaraki_ken','aizen_sou','urahara_ki','sh','gintama','sakata_gin','shimura_shi','kagura_(gintama)','shinseng','okita_sougo_(gintama)','hijikata_toush','kondo_isa','dragon_ball','son_goku','vegeta','bulma_(dragon_ball)','frieza','cell_(dragon_ball)','majin_buu','piccolo','gohan','trunks_(dragon_ball)','android_18','android_21'],
    '舰队collection': ['kantai_collection','kancolle','fubuki_(kancolle)','shimakaze_(kancolle)','kongou_(kancolle)','akagi_(kancolle)','kaga_(kancolle)','yamato_(kancolle)','nagato_(kancolle)','mutsu_(kancolle)','haruna_(kancolle)','kirishima_(kancolle)','hiei_(kancolle)','sendai_(kancolle)','jintsuu_(kancolle)','naka_(kancolle)','yuudachi_(kancolle)','shigure_(kancolle)','poi_(kancolle)','prinz_eugen_(kancolle)','bismarck_(kancolle)','ro_500_(kancolle)','iowa_(kancolle)','warspite_(kancolle)','saratoga_(kancolle)','gambier_bay_(kancolle)','taihou_(kancolle)','kashima_(kancolle)','z1_(kancolle)','z3_(kancolle)','ktkm_(kancolle)','ooyodo_(kancolle)'],
    '赛马娘': ['umamusume','pretty_derby','special_week','silence_suz','to','tokai_teio','mejiro_mcq','oguri_cap','gold_ship','daiwa_sca','vodka_(umamusume)','n','air_groove','kitasan_black','satono_diamond','twin_turbo','mejiro_palmer','manhattan_cafe','haru_urara','king_halo','grass_wonder','el_condor_pasa','tm_opera_o','narita_brian','symboli_rud','maruzensky','fuji_kise','agnes_tach','biwa_haya','inari_one','super_creek','sakura_baku','shinko_windy','sweep_tos','agnes_dig','dualingo_','jungle_poc','smart_falc','curren_cha','kawakami_p','copano_rickey','hokko_ta','wonder_acu','mejiro_dob','rice_shower','mejiro_ard','sakura_ch','sirius_sy','meisho'],
    '虚拟形象': ['virtual_youtuber','vtuber','kizuna_ai','kaguya_luna','mirai_akari','nojaloli','s','den','si','tenjin_kotom','tsukino_mito','higuchi_kae','shizuka_rin','achikita_ch','yashiro_ki','yuuki_chi','moira_(vtuber)','shinonom','virtual_streamer','vstreamer'],
    '主机游戏|掌机游戏|steam游戏': ['nier','nier_automata','2b_(nier)','9s_(nier)','a2_(nier)','ka','persona','persona_5','joker_(persona)','morgana_(persona)','makoto_ni','futaba_sa','ann_taka','yusuke_k','haru_oku','goro_ake','kasumi_yosh','persona_4','persona_3','shin_megami_tensei','smt','bayonetta','jeanne_(bayonetta)','fire_emblem','fe','super_smash_bros','smash_bros','legend_of_zelda','zelda','link_(zelda)','ganondorf','xenoblade','xenoblade_chronicles','pyra_(xenoblade)','mythra','splatoon','inkling','animal_crossing','monster_hunter','kirby_(series)','star_fox','metroid','samus_ar','final_fantasy','ff','tifa_lock','aerith_gains','lightning_(ff)','yuna_(ff)','kingdom_hearts','kh','sora_(kh)','dark_souls','elden_ring','bloodborne','sekiro','armored_core','ace_combat','resident_evil','re','jill_val','ada_wong','claire_red','street_fighter','chun_li','cammy_white','juri_han','tekken','guilty_gear','blazblue','under_night','melty_blood','tales_of','atelier','disgaea','danganronpa','steins_gate','ace_attorney','phoenix_wright','devil_may_cry','dmc','dante_(dmc)','vergil_(dmc)','nero_(dmc)','metal_gear','metal_gear_solid','mgs','snake_(mgs)','raiden_(mgs)','silent_hill','castlevania','vania'],
    '端游|页游|pc端': ['league_of_legends','lol','ahri_(lol)','akali_(lol)','ashe_(lol)','kai_sa','lux_(lol)','j','miss_fortune_(lol)','seraphine_(lol)','jinx_(lol)','vi_(lol)','caitlyn_(lol)','dva_(overwatch)','overwatch','tracer_(overwatch)','widowmaker_(overwatch)','mercy_(overwatch)','mei_(overwatch)','pharah_(overwatch)','brigitte_(overwatch)','ashe_(overwatch)','kiriko_(overwatch)','sombra_(overwatch)','symmetra_(overwatch)','ana_(overwatch)','valorant','jett_(valorant)','viper_(valorant)','sage_(valorant)','raze_(valorant)','reyna_(valorant)','killjoy_(valorant)','wow','world_of_warcraft','jaina_proud','sylvanas_wind','tyrande_whis','alexstrasza','ysera','onyxia','kerrigan','starctaft','diablo_(game)','lilith_(diablo)','minecraft','terraria','stardew_valley','skyrim','fallout','cyberpunk','cyberpunk_2077','the_witcher','geralt_of_rivia','yennefer','triss_merigold','ciri_(witcher)','baldurs_gate','bg3','shadowheart_(bg3)','lae_zel','karlach_(bg3)','astarion_(bg3)','gale_(bg3)','mass_effect','liara_t_soni','tali_zorah','miranda_lawson','jack_(mass_effect)','dragon_age','morrigan_(dragon_age)','leliana_(dragon_age)','isabela_(dragon_age)','merrill_(dragon_age)','borderlands','lilith_(borderlands)','maya_(borderlands)','gaige_(borderlands)','warframe','destiny_(game)','destiny_2','division_(game)','rainbow_six_siege','apex_legends','wraith_(apex)','loba_(apex)','wattson_(apex)','call_of_duty','pubg','fortnite','genshin','honkai','star_rail','wuthering','wuwa','maplestory','dungeon_fighter','dnf','lost_ark'],
    '动漫番剧': ['spy_x_family','yor_forger','anya_forger','chainsaw_man','csm','makima','power_(csm)','reze_(csm)','kobeni_higash','himeno_(csm)','asa_mit','fami_(csm)','yoru_(csm)','jujutsu_kaisen','jjk','gojo_satoru','itadori_yu','fushiguro_me','kugisaki_no','jujutsu','demon_slayer','kimetsu_no_ya','nezuko_kama','shinobu_koc','mitsuri_kan','daki_(demon_slayer)','tamayo_(demon_slayer)','kanroji_mits','attack_on_t','shingeki_no_ky','mikasa_ack','historia_re','annie_leon','sasha_blau','pieck_fin','hange_zoe','fullmetal_al','fma','winry_rock','riza_haw','lust_(fma)','one_punch_man','opm','tatsumaki','fubuki_(opm)','sword_art_online','sao','asuna_yu','sinon_(sao)','re_zero','emilia_(re_zero)','rem_(re_zero)','ram_(re_zero)','beatrice_(re_zero)','echidna_(re_zero)','konosuba','aqua_(konosuba)','megumin','darkness_(konosuba)','eris_(konosuba)','wiz_(konosuba)','overlord_(anime)','albedo_(overlord)','shalltear_blood','narberal_gamma','lupusregina_','evangelion','eva','ayanami_rei','asuka_langley_s','misato_kats','mari_makin','code_geass','cc_(code_geass)','kallen_stadt','euphemia_li_','shirley_fen','gurren_lagann','yoko_littner','darling_in_the_franxx','zero_two_(darling)','ichigo_(darling)','gundam','mobile_suit_gundam','macross','cowboy_bebop','faye_valen','ghost_in_the_shell','motoko_kusa','trigun','meryl_strife','mil]_thompson','hellsing','integra_hell','seras_victo','elfen_lied','lucy_(elfen_lied)','monster_musume','mia_(monster_musume)','centorea_shi','papi_(monster_musume)','suu_(monster_musume)','rachnera_ar','meroune_lore','konosuba','nichijou','k_on','yui_hiras','mio_akiyama','ritsu_tain','tsumugi_kot','azusa_nakano','lucky_star','haruhi_suzu','yuki_nagato','mikuru_asah','clannad','nagisa_fur','kyou_fujib','tomoyo_sak','toradora','taiga_ais','ami_kawash','angel_beats','kanade_tach','yuri_nakamura','guilty_crown','inori_yuzu','steins_gate','makise_kur','suzuha_amane','psycho_pass','akane_tsune','shougo_maki','violet_ever','violet_(violet_evergarden)','made_in_abyss','nanachi','ouzen','prushka_(made_in_abyss)','faputa_(made_in_abyss)','mushoku_tensei','roxy_migur','eris_boreas_g','sylphiette_(mushoku)','bocchi_the_rock','gotou_hito','kita_iku','yamada_ryo','ijichi_nij','frieren_(sousou_no_frieren)','fern_(frieren)','stark_(frieren)','aura_(frieren)','ubel_(frieren)','serie_(frieren)','flamme_(frieren)','mein_(frieren)','dungeon_meshi','marcille_do','falin_toud','lycion_(dungeon_meshi)','your_name','kimi_no_na_wa','mitsuha_miy','weathering_with_you','tenki_no_ko','suzume_no_to','suzume_iwa','spirited_away','sen_to_chihiro','princess_mononoke','howls_moving_castle','totoro','my_neighbor_totoro','kikis_delivery','nausicaa_(ghibli)','ghibli','ponyo','castle_in_the_sky','sheeta_(laputa)','dola_(laputa)'],
    '小众手游|小众二游': ['girls_band_cry','bang_dream','bandori','kasumi_toy','yukina_min','ran_miteke','aya_maruy','morfonica','project_sekai','hatsune_miku_(project_sekai)','kiritani_haru','azusawa_koh','yoisaki_kan','hinomori_sh','asahina_maf','hanasato_mi','tenma_tsu','otori_emu','kusanagi_ne','kamishiro_ru','aoyagi_tou','shinonome_a','shinonome_e','shiraishi_a','nene_kusa','roki_(vocaloid)','vocaloid','kagamine_rin','kagamine_len','megurine_lu','ia_(vocaloid)','gumi_(vocaloid)','kasane_teto','yowane_ha','akita_ne','utau','blue_reflection','ryza_(atelier)','atelier_ryza','klaudia_valentz','lila_decyrus','ar_tonelico','ar_nosurge','hyperdimension_neptunia','neptune_(neptunia)','noire_(neptunia)','blanc_(neptunia)','vert_(neptunia)','senran_kagura','asuka_(senran_kagura)','homura_(senran_kagura)','yumi_(senran_kagura)','ikaruga_(senran)','katsuragi_(senran)','to_love_ru','lala_satalin','momo_velia','nana_asta','yami_(to_love_ru)','haruna_(to_love_ru)','highschool_dxd','rias_gremory','akeno_hime','koneko_tou','asia_argent','xenoblade_2','homura_(xenoblade)','hikari_(xenoblade)','nia_(xenoblade)','zelda_botw','mipha','urbosa','riju_(zelda)','zelda_(botw)','paya_(zelda)','fire_emblem_three_houses','edelgard','dimitri_(fe)','claude_(fe)','byleth_(fe)','rhea_(fe)','lyndis','lyn_(fire_emblem)','lucina_(fe)','camilla_(fe)','azura_(fe)','corrin_(fe)','tharja_(fe)','honkai_impact','blue_archive','arknights','azur_lane','fgo','priconne','idolmaster','idoly_pride','lapis_relights','d4dj','aimoto_rinku','shirogane_to','azusawa_miy','love_live_superstar','liella','nijigasaki','aquors','muse_(love_live)','a_rise_(love_live)'],
}

def detect_ip(tag_name):
    """Detect which IP a tag belongs to"""
    tl = tag_name.lower()
    # Check copyright brackets
    bracket = re.search(r'\(([^)]+)\)', tag_name)
    bracket_text = bracket.group(1).lower() if bracket else ''
    
    for ip_name, keywords in IP_MAP.items():
        for kw in keywords:
            if kw == tl or kw in tl or kw in bracket_text:
                return ip_name
    return None

# ============================================
# STEP 4: Classify all tags
# ============================================
print("[3/6] Classifying hierarchical tags...")
classified = {}  # (tag_name, translate) -> [(cat1, cat2, hot), ...]

for tag_name, translate, sg_id in hier_tags:
    cat1, cat2 = map_hier_to_new(sg_id, tag_name)
    if cat1 is None:
        cat1, cat2 = '场景建筑', '其他场景'
    
    # Handle hair splitting
    if cat1 == '头发头部':
        cat2 = classify_hair_tag(tag_name)
    
    hot = danbooru_tags.get(tag_name, ('', 0))[1]
    hot_str = f'{hot:,}' if hot else '—'
    classified[(tag_name, translate)] = [(cat1, cat2, hot_str)]
    
    # Cross-category: ears/精灵/兽人 → also 角色tag
    if cat1 == '种族生物' and cat2 in ['耳朵', '精灵/妖精', '兽人/亚人', '天使/恶魔', '龙族', '妖怪/怪物']:
        if cat2 == '耳朵':
            classified[(tag_name, translate)].append(('角色tag', '精灵角色' if 'elf' in tag_name.lower() else '兽人/亚人角色', hot_str))
        elif cat2 == '精灵/妖精':
            classified[(tag_name, translate)].append(('角色tag', '精灵角色', hot_str))
        elif cat2 == '兽人/亚人':
            classified[(tag_name, translate)].append(('角色tag', '兽人/亚人角色', hot_str))
        else:
            classified[(tag_name, translate)].append(('角色tag', '其他角色', hot_str))

print(f"  Classified {len(classified)} unique tags from hierarchy")

print("[4/6] Classifying danbooru-only tags...")
# Build lookup from classified tags
classified_tagnames = set(tn for tn, _ in classified.keys())

dn_only_count = 0
for tag_name, (translate, hot) in danbooru_tags.items():
    if tag_name in classified_tagnames:
        continue
    
    hot_str = f'{hot:,}' if hot else '—'
    tl = tag_name.lower()
    
    # Rule-based classification
    classified_here = False
    
    # Try NSFW
    nsfw_cat = classify_nsfw(tag_name)
    if nsfw_cat:
        classified[(tag_name, translate)] = [('nsfw内容', nsfw_cat, hot_str)]
        classified_here = True
        continue
    
    # Try IP detection for character tags
    ip = detect_ip(tag_name)
    if ip:
        classified[(tag_name, translate)] = [('角色tag', ip, hot_str)]
        classified_here = True
        continue
    
    # Clothing
    if any(kw in tl for kw in ['shirt','dress','skirt','pants','shorts','jeans','jacket','coat','sweater','hoodie','t_shirt','tank_top','blouse','uniform','suit','kimono','yukata','vest','cardigan','poncho','cape','robe','cloak','apron','swimsuit','bikini','leotard','bodysuit','jumpsuit','romper','overalls','sweatshirt','sweatpants','leggings','stockings','tights','socks','pantyhose','loincloth','fundoshi','sarong','hakama','hanfu','qipao','cheongsam','sailor_uniform','school_uniform','military_uniform','nurse_uniform','maid_uniform','police_uniform','flight_attendant_uniform','sportswear','armor','cuirass','gauntlets','greaves','pauldrons','helmet','chainmail','plate_armor','leather_armor','robe','belt','suspenders','garter','garter_belt','corset','bustier','brassiere','bra','panties','boxers','briefs','thong','fundoshi','loincloth','sarashi','bandeau','tube_top','camisole','negligee','lingerie','underwear','footwear','boots','shoes','sandals','slippers','heels','high_heels','stilettos','loafers','sneakers','flip_flops','geta','tabi','zori','wooden_sandals']):
        classified[(tag_name, translate)] = [('服装', '其他服装', hot_str)]
        classified_here = True
        continue
    
    # Hair
    if 'hair' in tl or 'wig' in tl or 'bald' in tl or 'bangs' in tl or 'ponytail' in tl or 'braid' in tl or 'twintails' in tl or 'bun' in tl:
        subcat = classify_hair_tag(tag_name)
        classified[(tag_name, translate)] = [('头发头部', subcat, hot_str)]
        classified_here = True
        continue
    
    # Face
    if any(kw in tl for kw in ['eye','eyes','eyebrow','eyelash','eyelid','pupil','nose','mouth','lip','lips','tongue','teeth','tooth','facial','blush']):
        classified[(tag_name, translate)] = [('脸部元素', '面部', hot_str)]
        classified_here = True
        continue
    
    # Body
    if any(kw in tl for kw in ['breast','nipple','cleavage','waist','hip','thigh','leg','arm','hand','finger','foot','feet','belly','navel','stomach','abs','muscle','skin','tan','tan_line','scar','tattoo','freckles','mole','beauty_mark','stretch_marks','cellulite','curvy','plump','chubby','petite','tall','short','height','body','figure','physique']):
        # But not NSFW body parts (those caught earlier)
        classified[(tag_name, translate)] = [('身体', '其他身体', hot_str)]
        classified_here = True
        continue
    
    # Expression
    if any(kw in tl for kw in ['smile','frown','cry','tears','angry','mad','happy','sad','surprise','shock','blush','embarrass','grin','laugh','pout','scowl','wince','scream','yawn','sleepy','tired','bored','serious','nervous','scared','afraid','disgust','pain','expression','emotion']):
        classified[(tag_name, translate)] = [('表情情绪', '其他表情', hot_str)]
        classified_here = True
        continue
    
    # Action
    if any(kw in tl for kw in ['sitting','standing','lying','walking','running','jumping','flying','falling','kneeling','crouching','bending','leaning','stretching','reaching','holding','grabbing','carrying','pulling','pushing','lifting','throwing','catching','pointing','waving','clapping','touching','hugging','kissing','dancing','fighting','punching','kicking','blocking','dodging','swinging','riding','driving','eating','drinking','cooking','reading','writing','drawing','painting','singing','playing','sleeping','waking','bathing','swimming','diving','climbing','crawling']):
        classified[(tag_name, translate)] = [('动作', '其他动作', hot_str)]
        classified_here = True
        continue
    
    # Background
    if 'background' in tl:
        classified[(tag_name, translate)] = [('背景', '背景', hot_str)]
        classified_here = True
        continue
    
    # Lighting
    if any(kw in tl for kw in ['light','lighting','shadow','glow','shiny','lens_flare','backlight','rim_light','sunlight','moonlight','spotlight','dappled_light','god_rays','crepuscular_rays','reflection','bloom','glare']):
        classified[(tag_name, translate)] = [('光照风格', '光照', hot_str)]
        classified_here = True
        continue
    
    # Art style
    if any(kw in tl for kw in ['style','sketch','lineart','watercolor','oil_painting','painting','render','cg','3d','2d','pixel','monochrome','greyscale','grayscale','sepia','colored','flat_color','gradient','brush','pencil','ink','marker','crayon','charcoal','pastel','airbrush','digital','traditional','photorealistic','realistic','cartoon','anime','manga','comic','illustration','vector']):
        classified[(tag_name, translate)] = [('绘图风格', '其他风格', hot_str)]
        classified_here = True
        continue
    
    # Scene
    if any(kw in tl for kw in ['room','house','building','school','classroom','office','library','hospital','church','temple','shrine','castle','tower','bridge','street','road','alley','park','garden','forest','beach','ocean','sea','river','lake','mountain','hill','field','desert','snow','rain','cloud','sky','sunset','sunrise','night','day','city','town','village','ruins','door','window','wall','floor','ceiling','roof','balcony','stairs','hallway','corridor','bathroom','kitchen','bedroom','living_room','cafe','restaurant','shop','store','market','stadium','arena','stadium','pool','onsen','bath','shower','toilet','mirror','sofa','couch','chair','table','desk','bed','pillow','blanket','curtain','carpet','lamp','chandelier','candle','fireplace','book','bookshelf','bookcase']):
        classified[(tag_name, translate)] = [('场景建筑', '其他场景', hot_str)]
        classified_here = True
        continue
    
    # Accessories
    if any(kw in tl for kw in ['necklace','bracelet','ring','anklet','brooch','pendant','choker','chain','watch','cufflinks','earring','piercing','glasses','sunglasses','goggles','monocle','mask','scarf','gloves','bag','purse','backpack','umbrella','parasol','cane','fan','ribbon','bow','bell','collar','leash','wings','tail','halo','horn']):
        classified[(tag_name, translate)] = [('配件配饰', '其他配饰', hot_str)]
        classified_here = True
        continue
    
    # Age/occupation
    if any(kw in tl for kw in ['loli','shota','child','kid','baby','toddler','teen','teenager','adult','elder','old','young','age','nurse','doctor','teacher','student','maid','butler','police','officer','soldier','knight','queen','king','princess','prince','witch','wizard','mage','priest','nun','chef','waitress','idol','singer','dancer','artist','model','office_lady','salaryman','delinquent','yakuza','samurai','ninja','pirate','cowboy','cowgirl','angel','devil','god','goddess','reaper','vampire','werewolf','zombie','ghost','slime','fairy']):
        classified[(tag_name, translate)] = [('年龄职业', '身份', hot_str)]
        classified_here = True
        continue
    
    # Number of people
    if re.match(r'^(\d+)(girl|boy|girls|boys|people|person)$', tl) or tl in ['solo','multiple_girls','multiple_boys','group','crowd','no_humans']:
        classified[(tag_name, translate)] = [('年龄职业', '人数', hot_str)]
        classified_here = True
        continue
    
    # Creature
    if any(kw in tl for kw in ['animal','creature','monster','dragon','beast','kaiju','robot','mecha','alien','slime','golem','gargoyle','centaur','mermaid','lamia','harpy','siren','succubus','incubus','oni','kitsune','tanuki','nekomata','bakeneko','yokai','youkai','fairy','pixie','elf','dwarf','orc','goblin','troll','ogre','giant','minotaur','cyclops','medusa','gorgon','sphinx','griffin','phoenix','chimera','hydra','cerberus','kraken','werewolf','vampire','zombie','skeleton','lich','ghost','spirit','elemental']):
        classified[(tag_name, translate)] = [('种族生物', '幻想生物/其他', hot_str)]
        classified_here = True
        continue
    
    # Plants
    if any(kw in tl for kw in ['flower','plant','tree','grass','leaf','rose','lily','sunflower','cherry_blossom','sakura','bamboo','vine','mushroom','fungus','cactus','bush','forest','garden']):
        classified[(tag_name, translate)] = [('种族生物', '植物', hot_str)]
        classified_here = True
        continue
    
    # Composition
    if any(kw in tl for kw in ['close_up','closeup','extreme_closeup','medium_shot','long_shot','full_body','upper_body','lower_body','headshot','portrait','half_body','cowboy_shot','dutch_angle','worm_eye','bird_eye','birds_eye','low_angle','high_angle','aerial','panoramic','fisheye','wide_angle','telephoto','depth_of_field','bokeh','blurred_background','blurry_background','focus','framed','frame','border','vignette','letterbox','cinematic','wide_screen','over_the_shoulder','from_behind','from_above','from_below','side_view','front_view','back_view','profile','three_quarter','isometric','perspective','pov','point_of_view','first_person','selfie','mirror_selfie','webcam','photograph','photo','snapshot','selfie']):
        classified[(tag_name, translate)] = [('视角构图', '其他构图', hot_str)]
        classified_here = True
        continue
    
    # If nothing matched → 角色tag.其他角色 (most likely a character name)
    classified[(tag_name, translate)] = [('角色tag', '其他角色', hot_str)]
    dn_only_count += 1

print(f"  Classified {dn_only_count} danbooru-only tags")

# ============================================
# STEP 5: Build output rows
# ============================================
print("[5/6] Building output rows...")
output_rows = []
for (tag_name, translate), cat_list in classified.items():
    for cat1, cat2, hot_str in cat_list:
        output_rows.append([translate, tag_name, cat1, cat2, hot_str])

# Sort by category order
cat_rank = {c: i for i, c in enumerate(CAT_ORDER)}
output_rows.sort(key=lambda r: (cat_rank.get(r[2], 99), r[3], r[0]))

print(f"  Total output rows: {len(output_rows)}")

# ============================================
# STEP 6: Generate Excel
# ============================================
print("[6/6] Generating Excel...")

wb = openpyxl.Workbook()

# ---- Sheet 1: 全量标签对照表 ----
ws = wb.active
ws.title = "全量标签对照表"

header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_font = Font(bold=True, size=11, color='FFFFFF', name='微软雅黑')
data_font = Font(size=10, name='微软雅黑')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

cat_fills = {
    '服装': 'E2EFDA', '种族生物': 'FCE4D6', '角色tag': 'D9E2F3',
    '表情情绪': 'FFF2CC', '动作': 'E4DFEC', 'nsfw内容': 'F4B4C2',
    '配件配饰': 'DDEBF7', '脸部元素': 'E2EFDA', '身体': 'FCE4D6',
    '头发头部': 'FCE4D6', '场景建筑': 'FFF2CC', '年龄职业': 'E4DFEC',
    '绘图风格': 'DDEBF7', '背景': 'F2F2F2', '光照风格': 'FFFFCC',
    '视角构图': 'F2F2F2',
}

headers = ['中文翻译', '标签名(tag)', '一级分类', '二级分类', '热度']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

for i, row in enumerate(output_rows, 2):
    cat = row[2]
    color = cat_fills.get(cat)
    row_fill = PatternFill(start_color=color, end_color=color, fill_type='solid') if color else None
    for j, val in enumerate(row, 1):
        cell = ws.cell(row=i, column=j, value=val)
        cell.font = data_font
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        if j == 5:
            cell.alignment = Alignment(horizontal='right')
        elif j in [3, 4]:
            cell.alignment = Alignment(horizontal='center')
    if i % 10000 == 0:
        print(f"    Row {i}/{len(output_rows)}...")

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 34
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 14
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f"A1:E{len(output_rows)+1}"

# ---- Sheet 2: 分类结构概览 ----
ws2 = wb.create_sheet("分类结构概览")
ws2.cell(row=1, column=1, value="一级分类").font = Font(bold=True, name='微软雅黑')
ws2.cell(row=1, column=2, value="二级分类").font = Font(bold=True, name='微软雅黑')
ws2.cell(row=1, column=3, value="标签数量").font = Font(bold=True, name='微软雅黑')
ws2.cell(row=1, column=4, value="说明").font = Font(bold=True, name='微软雅黑')
for c in range(1, 5):
    ws2.cell(row=1, column=c).fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    ws2.cell(row=1, column=c).font = Font(bold=True, color='FFFFFF', name='微软雅黑')

# Count tags per category+subcategory
from collections import Counter
cat_count = Counter()
subcat_count = Counter()
for row in output_rows:
    cat_count[row[2]] += 1
    subcat_count[(row[2], row[3])] += 1

row_idx = 2
descriptions = {
    '角色tag-二次元角色(通用)': '非特定IP的角色标签',
    '角色tag-精灵角色': '跨类标签(同种族生物.精灵/妖精)',
    '角色tag-兽人/亚人角色': '跨类标签(同种族生物.兽人/亚人)',
    '角色tag-其他角色': '无法确定IP的兜底角色',
    '绘图风格-魔法特效': '原自定义(魔法1.0/魔法1.5)并入',
    '服装-汉服-唐风': '汉服子分类合并',
    '服装-汉服-宋风': '汉服子分类合并',
    '服装-汉服-明风': '汉服子分类合并',
    '服装-汉服-其他': '汉服子分类合并兜底',
}

for cat in CAT_ORDER:
    subcats = CATEGORY_CONFIG[cat]
    total = cat_count.get(cat, 0)
    ws2.cell(row=row_idx, column=1, value=cat).font = Font(bold=True, size=11, name='微软雅黑')
    ws2.cell(row=row_idx, column=3, value=total).font = Font(bold=True, name='微软雅黑')
    ws2.cell(row=row_idx, column=3).alignment = Alignment(horizontal='right')
    for sc in subcats:
        cnt = subcat_count.get((cat, sc), 0)
        desc = descriptions.get(f'{cat}-{sc}', '')
        ws2.cell(row=row_idx, column=2, value=sc).font = Font(name='微软雅黑')
        ws2.cell(row=row_idx, column=3, value=cnt).font = Font(name='微软雅黑')
        ws2.cell(row=row_idx, column=3).alignment = Alignment(horizontal='right')
        ws2.cell(row=row_idx, column=4, value=desc).font = Font(size=9, color='666666', name='微软雅黑')
        row_idx += 1

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 36
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 42
ws2.freeze_panes = 'A2'

# ---- Sheet 3: 统计 ----
ws3 = wb.create_sheet("统计")
ws3.cell(row=1, column=1, value="一级分类").font = Font(bold=True, name='微软雅黑')
ws3.cell(row=1, column=2, value="标签行数").font = Font(bold=True, name='微软雅黑')
ws3.cell(row=1, column=3, value="占比").font = Font(bold=True, name='微软雅黑')
for c in range(1,4):
    ws3.cell(row=1, column=c).fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    ws3.cell(row=1, column=c).font = Font(bold=True, color='FFFFFF', name='微软雅黑')

total_rows = len(output_rows)
for i, cat in enumerate(CAT_ORDER, 2):
    cnt = cat_count.get(cat, 0)
    ws3.cell(row=i, column=1, value=cat).font = Font(name='微软雅黑')
    ws3.cell(row=i, column=2, value=cnt).font = Font(name='微软雅黑')
    ws3.cell(row=i, column=2).alignment = Alignment(horizontal='right')
    ws3.cell(row=i, column=3, value=f'{cnt/total_rows*100:.1f}%').font = Font(name='微软雅黑')
    ws3.cell(row=i, column=3).alignment = Alignment(horizontal='right')

ws3.cell(row=i+1, column=1, value="合计").font = Font(bold=True, name='微软雅黑')
ws3.cell(row=i+1, column=2, value=total_rows).font = Font(bold=True, name='微软雅黑')
ws3.cell(row=i+1, column=2).alignment = Alignment(horizontal='right')
ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 14
ws3.column_dimensions['C'].width = 12

out_path = r"D:\AIANDshezhi\GenericAgent\temp\标签分类对照表_全量.xlsx"
wb.save(out_path)
print(f"\n{'='*60}")
print(f"✅ 全量Excel已生成: {out_path}")
print(f"   总行数: {total_rows:,}")
print(f"   Sheet 1: 全量标签对照表")
print(f"   Sheet 2: 分类结构概览")
print(f"   Sheet 3: 统计")
print(f"{'='*60}")

# Print distribution
print("\n=== 分布 ===")
for cat in CAT_ORDER:
    cnt = cat_count.get(cat, 0)
    print(f"  {cat}: {cnt:,} 行 ({cnt/total_rows*100:.1f}%)")
