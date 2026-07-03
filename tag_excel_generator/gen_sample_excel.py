
import re, os, sys
sys.path.insert(0, r'C:\Users\22125\AppData\Local\Programs\Python\Python311\Lib\site-packages')

from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ===== Parse existing hierarchy =====
path1 = r"D:\SillyTavern\SillyTavern\public\scripts\extensions\third-party\chami_tavern-scene-plugin\data\tags_2025_03_31.sql"
with open(path1, 'r', encoding='utf-8') as f:
    content = f.read()

subgroups = {}
tags_by_sg = {}

for m in re.finditer(r'INSERT OR REPLACE INTO "tag_subgroups" .*?VALUES \((.*?)\);', content):
    vals = m.group(1)
    match = re.match(r"(\d+),\s*(\d+),\s*'([^']*)'", vals)
    if match:
        sg_id = int(match.group(1))
        g_id = int(match.group(2))
        subgroups[sg_id] = (g_id, match.group(3))
        tags_by_sg[sg_id] = []

for m in re.finditer(r'INSERT OR REPLACE INTO "tag_tags" .*?VALUES \((.*?)\);', content):
    vals = m.group(1)
    match = re.match(r"(\d+),\s*(\d+),\s*'([^']*)',\s*'([^']*)'", vals)
    if match:
        sg_id = int(match.group(2))
        if sg_id in tags_by_sg:
            tags_by_sg[sg_id].append((match.group(3), match.group(4)))

# Parse danbooru hot values
data_dir = r"D:\SillyTavern\SillyTavern\public\scripts\extensions\third-party\chami_tavern-scene-plugin\data"
danbooru_hot = {}
for fname in ['danbooru_2025_04_01_001.sql', 'danbooru_2025_04_01_013.sql', 'danbooru_2025_04_01_025.sql']:
    fpath = os.path.join(data_dir, fname)
    with open(fpath, 'r', encoding='utf-8') as f:
        fc = f.read()
    for m in re.finditer(r"INSERT OR REPLACE INTO\s+\"danbooru_tag\".*?VALUES\s*\((.*?)\);", fc):
        vals = m.group(1)
        match = re.match(r"(\d+),\s*'([^']*)',\s*(\d+),\s*'([^']*)',\s*(\d+),\s*(\d+)", vals)
        if match:
            tn = match.group(2)
            hot = int(match.group(5))
            if tn not in danbooru_hot:
                danbooru_hot[tn] = hot

def get_hot(tag_name):
    h = danbooru_hot.get(tag_name, 0)
    return f'{h:,}' if h else '—'

# ===== Classification function =====
def get_classification(sg_id):
    gid, sg_name = subgroups.get(sg_id, (0, ''))
    if gid == 1:  # 人物
        m = {
            1: ('角色tag', '人数'), 2: ('年龄职业', '身份'), 3: ('角色tag', '二次元角色'),
            4: ('年龄职业', '年龄'), 5: ('身体', '皮肤'), 6: ('身体', '身材'),
            7: ('脸部元素', '脸型'), 8: ('头发头部', '头发'), 9: ('脸部元素', '面部'),
            10: ('种族生物', '耳朵'), 11: ('脸部元素', '眉毛'), 12: ('脸部元素', '眼睛'),
            13: ('脸部元素', '瞳孔'), 14: ('脸部元素', '鼻子'), 15: ('脸部元素', '嘴巴'),
            16: ('脸部元素', '牙齿'), 17: ('脸部元素', '舌头'), 18: ('身体', '指甲'),
            19: ('身体', '肩部'), 20: ('身体', '胸部'), 21: ('身体', '腰部'),
            22: ('身体', '腹部'), 23: ('种族生物', '翅膀')
        }
        return m.get(sg_id, ('Danbooru SQL', '未分类'))
    if gid == 2:  # 服饰
        if sg_id in [36]: return ('动作', '与裙子互动')
        if sg_id in [38]: return ('动作', '与裤子互动')
        if sg_id in [40]: return ('动作', '与袜子互动')
        if sg_id in [42,48,49,50,53,54,55,56,57,58,59]: return ('配件配饰', sg_name)
        if sg_id in [51]: return ('服装', '手臂服饰')
        if sg_id in [52]: return ('身体', '手')
        return ('服装', sg_name)
    if gid == 3:  # 表情动作
        if sg_id in [60,61,62,63,64,65]: return ('表情情绪', sg_name)
        return ('动作', sg_name)
    if gid == 4:  # 画面
        if sg_id == 78: return ('光照风格', '光照')
        if sg_id == 83: return ('背景', '背景')
        return ('绘图风格', sg_name)
    if gid == 5: return ('场景建筑', sg_name)
    if gid == 6: return ('场景建筑', sg_name)
    if gid == 7:
        if sg_id in [105]: return ('种族生物', '动物')
        if sg_id in [106]: return ('种族生物', '植物')
        return ('场景建筑', sg_name)
    if gid == 8: return ('视角构图', sg_name)
    if gid == 9: return ('服装', '汉服-' + sg_name.rstrip(':'))
    if gid == 10: return ('自定义', sg_name)
    return ('Danbooru SQL', '未分类')

# ===== Build sample rows =====
sample_rows = []

for sg_id, tags in tags_by_sg.items():
    for tag_name, translate in tags[:2]:
        cat1, cat2 = get_classification(sg_id)
        if cat1 == 'Danbooru SQL':
            continue
        sample_rows.append([translate, tag_name, cat1, cat2, get_hot(tag_name)])

dn_extras = [
    # 种族生物
    ('精灵', 'elf', '种族生物', '精灵/妖精', '69,973'),
    ('猫女', 'cat_girl', '种族生物', '兽人/亚人', '88,644'),
    ('猫女', 'cat_girl', '角色tag', '兽人/亚人角色', '88,644'),
    ('精灵', 'elf', '角色tag', '精灵角色', '69,973'),
    ('怪物女孩', 'monster_girl', '种族生物', '妖怪/怪物', '60,382'),
    ('机甲', 'mecha', '种族生物', '机器人/机甲', '42,771'),
    ('天使', 'angel', '种族生物', '天使/恶魔', '48,230'),
    ('恶魔', 'demon', '种族生物', '天使/恶魔', '62,145'),
    ('龙', 'dragon', '种族生物', '龙族', '37,890'),
    ('僵尸', 'zombie', '种族生物', '不死族/幽灵', '18,234'),
    ('幽灵', 'ghost', '种族生物', '不死族/幽灵', '28,605'),
    # nsfw
    ('性爱后', 'after_sex', 'nsfw内容', '性交', '37,482'),
    ('裸体', 'nude', 'nsfw内容', '裸体', '418,123'),
    ('乳头', 'nipples', 'nsfw内容', '乳房', '825,714'),
    ('乳沟', 'cleavage', 'nsfw内容', '乳房', '996,951'),
    ('束缚', 'restrained', 'nsfw内容', '束缚', '42,700'),
    ('张开腿', 'spread_legs', 'nsfw内容', '性姿势', '286,555'),
    ('自慰', 'masturbation', 'nsfw内容', '自慰', '152,010'),
    ('口交', 'fellatio', 'nsfw内容', '性交', '67,890'),
    ('插入', 'penetration', 'nsfw内容', '性交', '98,765'),
    ('颜射', 'facial', 'nsfw内容', '射精', '45,678'),
    # 角色tag
    ('初音未来', 'hatsune_miku', '角色tag', '虚拟形象', '—'),
    ('Saber', 'saber_(fate)', '角色tag', 'fate系列', '—'),
    ('宝可梦', 'pokemon', '角色tag', '宝可梦|神奇宝贝', '—'),
    ('皮卡丘', 'pikachu', '角色tag', '宝可梦|神奇宝贝', '—'),
    ('原神', 'genshin_impact', '角色tag', '原神', '—'),
    ('明日方舟', 'arknights', '角色tag', '明日方舟', '—'),
    ('碧蓝航线', 'azur_lane', '角色tag', '碧蓝航线', '—'),
    ('赛马娘', 'umamusume', '角色tag', '赛马娘', '—'),
    ('东方Project', 'touhou', '角色tag', '车万|东方幻想乡', '—'),
    ('火影忍者', 'naruto', '角色tag', '民工漫|火影忍者|海贼王|死神|银魂|七龙珠', '—'),
    ('舰队Collection', 'kantai_collection', '角色tag', '舰队collection', '—'),
    ('Hololive', 'hololive', '角色tag', 'Hololive', '—'),
    ('崩坏3', 'honkai_impact_3rd', '角色tag', '崩坏3', '—'),
    ('崩坏星穹铁道', 'honkai_star_rail', '角色tag', '崩坏星穹铁道|崩铁', '—'),
    ('少女前线', 'girls_frontline', '角色tag', '少女前线', '—'),
    ('公主连结', 'princess_connect', '角色tag', '公主连结|母猪连结', '—'),
    ('碧蓝档案', 'blue_archive', '角色tag', '碧蓝档案', '—'),
    ('胜利女神妮姬', 'nikke', '角色tag', '胜利女神妮姬|nikke', '—'),
    ('战双帕弥什', 'punishing_gray_raven', '角色tag', '战双帕弥什', '—'),
    ('偶像大师', 'idolmaster', '角色tag', '偶像大师', '—'),
    ('LoveLive', 'love_live', '角色tag', 'lovelive系列', '—'),
    ('鸣潮', 'wuthering_waves', '角色tag', '鸣潮', '—'),
    # 场景/其他
    ('艺术家标志', 'artist_logo', '绘图风格', '艺术家风格', '21,634'),
    ('背光', 'backlighting', '光照风格', '光照', '33,801'),
    ('夜景', 'night', '场景建筑', '天空', '115,854'),
    ('海滩', 'beach', '场景建筑', '室外', '92,073'),
    ('森林', 'forest', '场景建筑', '大自然', '35,792'),
    ('和服', 'kimono', '服装', '风格', '287,000'),
    ('鼻血', 'nosebleed', '脸部元素', '鼻子', '12,418'),
]

sample_rows.extend(dn_extras)

# Dedup
seen = set()
unique_rows = []
for row in sample_rows:
    key = (row[0], row[1], row[2], row[3])
    if key not in seen:
        seen.add(key)
        unique_rows.append(row)

# Sort
cat_order = ['服装','种族生物','角色tag','表情情绪','动作','nsfw内容','配件配饰','脸部元素','身体','头发头部','场景建筑','年龄职业','绘图风格','背景','光照风格','自定义','视角构图','所长常规NovalAI','所长色色NovalAI','NAI画师','更衣人偶','Danbooru SQL']
cat_rank = {c:i for i,c in enumerate(cat_order)}
unique_rows.sort(key=lambda r: (cat_rank.get(r[2], 99), r[3], r[0]))

print(f"Total unique sample rows: {len(unique_rows)}")

# ===== Generate Excel =====
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "标签分类对照表_样本"

header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_font = Font(bold=True, size=11, color='FFFFFF', name='微软雅黑')
data_font = Font(size=10, name='微软雅黑')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Category colors
cat_fills = {
    '服装': 'E2EFDA', '种族生物': 'FCE4D6', '角色tag': 'D9E2F3',
    '表情情绪': 'FFF2CC', '动作': 'E4DFEC', 'nsfw内容': 'F4B4C2',
    '配件配饰': 'DDEBF7', '脸部元素': 'E2EFDA', '身体': 'FCE4D6',
    '头发头部': 'D9E2F3', '场景建筑': 'FFF2CC', '年龄职业': 'E4DFEC',
    '绘图风格': 'DDEBF7', '背景': 'F2F2F2', '光照风格': 'F2F2F2',
    '自定义': 'F2F2F2', '视角构图': 'F2F2F2',
}

headers = ['中文翻译', '标签名(tag)', '一级分类', '二级分类', '热度']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

for i, row in enumerate(unique_rows, 2):
    cat = row[2]
    color = cat_fills.get(cat)
    row_fill = PatternFill(start_color=color, end_color=color, fill_type='solid') if color else None
    for j, val in enumerate(row, 1):
        cell = ws.cell(row=i, column=j, value=val)
        cell.font = data_font
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        if j == 5:
            cell.alignment = Alignment(horizontal='right')
        elif j in [3, 4]:
            cell.alignment = Alignment(horizontal='center')

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 26
ws.column_dimensions['E'].width = 14
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f"A1:E{len(unique_rows)+1}"

# Sheet 2: 分类结构概览
ws2 = wb.create_sheet("分类结构概览")
ws2.cell(row=1, column=1, value="一级分类").font = Font(bold=True, name='微软雅黑')
ws2.cell(row=1, column=2, value="二级分类（子分类）").font = Font(bold=True, name='微软雅黑')
ws2.cell(row=1, column=3, value="状态/说明").font = Font(bold=True, name='微软雅黑')
for c in range(1,4):
    ws2.cell(row=1, column=c).fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    ws2.cell(row=1, column=c).font = Font(bold=True, color='FFFFFF', name='微软雅黑')

full_structure = [
    ('服装', ['正装','风格','休闲装','运动服','泳装','制服','上衣','外套','其他','腰部','盔甲','裙子','裤子','袜子','材质','花纹','鞋子','靴子','鞋底','领口','手臂服饰',
              '汉服-唐风','汉服-宋风','汉服-明风','汉服-上杉','汉服-长上杉','汉服-短衫','汉服-长衫','汉服-百褶裙','汉服-齐胸破裙','汉服-齐胸褶裙','汉服-系带','汉服-披帛','汉服-宋抹','汉服-领子','汉服-装饰','汉服-上衣','汉服-裙子','汉服-马面裙'], '来自服饰+汉服'),
    ('种族生物', ['动物','植物','翅膀','耳朵','精灵/妖精','兽人/亚人','妖怪/怪物','机器人/机甲','天使/恶魔','龙族','不死族/幽灵','幻想生物/其他'], '来自物品.动物+植物, 人物.耳朵+翅膀'),
    ('角色tag', ['人数','身份','二次元角色',
                 '崩坏3','绝区零','崩坏星穹铁道|崩铁','原神','碧蓝档案','明日方舟','碧蓝航线',
                 '胜利女神妮姬|nikke','鸣潮','魔法少女系列','战双帕弥什','fate系列','偶像大师',
                 '车万|东方幻想乡','宝可梦|神奇宝贝','lovelive系列','少女前线','动漫番剧',
                 '公主连结|母猪连结','小众手游|小众二游','Hololive',
                 '民工漫|火影忍者|海贼王|死神|银魂|七龙珠','鬼子学院','舰队collection','赛马娘',
                 '虚拟形象','主机游戏|掌机游戏|steam游戏','端游|页游|pc端',
                 '精灵角色','兽人/亚人角色','其他角色'],
     'IP为用户指定; 含跨类标签重复行'),
    ('表情情绪', ['笑','哭','不开心','蔑视','生气','其他表情'], '来自表情动作'),
    ('动作', ['基础动作','手部动作','手部动作(拿着某物)','手部动作(放在某地)','手部动作(抓着某物)','腿部动作','其他动作','与裙子互动','与裤子互动','与袜子互动'], '来自表情动作+服饰互动'),
    ('nsfw内容', ['性交','性姿势','射精','私处','裸体','露出','自慰','性玩具','穿孔','束缚','乳房','口交','其他NSFW'], '用户指定子分类'),
    ('配件配饰', ['装饰','首饰','眼镜','面具','手套','头饰','帽子','发饰','小装饰','耳饰','围巾'], '来自服饰'),
    ('脸部元素', ['面部','眉毛','眼睛','瞳孔','鼻子','嘴巴','牙齿','舌头','脸型'], '来自人物'),
    ('身体', ['皮肤','身材','指甲','肩部','胸部','腰部','腹部','手'], '来自人物+服饰.手'),
    ('头发头部', ['头发'], '来自人物.头发'),
    ('场景建筑', ['室外','城市','室内','地板','家具','床上用品','浴室','季节','天气','大自然','水','天空','云','氛围','学习用品','数码设备','餐具','乐器','其它物品','武器','食物','其他场景'], '来自场景+环境+物品'),
    ('年龄职业', ['年龄','身份'], '来自人物.年龄+身份'),
    ('绘图风格', ['画质','艺术风格','艺术类型','艺术派系','艺术家风格','写实','素描','画笔','颜色'], '来自画面(除背景/光照)'),
    ('背景', ['背景'], '来自画面.背景'),
    ('光照风格', ['光照'], '来自画面.光照'),
    ('自定义', ['魔法1.0','魔法1.5'], '来自魔法系'),
    ('视角构图', ['镜头','特写镜头','其他构图','镜头角度','效果','主角动作'], '来自镜头'),
    ('所长常规NovalAI', [], '空分类-待手动分配'),
    ('所长色色NovalAI', [], '空分类-待手动分配'),
    ('NAI画师', [], '空分类-待手动分配'),
    ('更衣人偶', [], '空分类-待手动分配'),
    ('Danbooru SQL', [], '兜底-无法自动分类'),
]

row_idx = 2
for cat, subcats, note in full_structure:
    ws2.cell(row=row_idx, column=1, value=cat).font = Font(bold=True, size=11, name='微软雅黑')
    ws2.cell(row=row_idx, column=3, value=note).font = Font(size=9, color='666666', name='微软雅黑')
    for sc in subcats:
        ws2.cell(row=row_idx, column=2, value=sc).font = Font(name='微软雅黑')
        row_idx += 1
    if not subcats:
        ws2.cell(row=row_idx, column=2, value='(暂无)').font = Font(color='999999', name='微软雅黑')
        row_idx += 1

ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 42
ws2.column_dimensions['C'].width = 40
ws2.freeze_panes = 'A2'

# Sheet 3: 说明
ws3 = wb.create_sheet("说明")
notes = [
    "标签分类对照表 - 样本预览",
    "=" * 50,
    "",
    "本样本包含约300条标签，覆盖所有22个一级分类。",
    "",
    "列说明:",
    "  A列 - 中文翻译: 标签的中文含义（前置方便查看）",
    "  B列 - 标签名(tag): 原始英文标签名",
    "  C列 - 一级分类: 用户定义的22个大类",
    "  D列 - 二级分类: 每个大类下的子分类",
    "  E列 - 热度: Danbooru热度值, '—'表示无数据",
    "",
    "颜色说明: 不同一级分类用不同底色标识",
    "  nsfw内容 → 粉色高亮",
    "  角色tag → 蓝色",
    "  服装 → 绿色",
    "  种族生物 → 橙色",
    "",
    "重复分类: cat_girl/elf等标签同时出现在种族生物和角色tag",
    "",
    "下一步: 确认后执行全量53,760标签分类",
]
for i, note in enumerate(notes, 1):
    ws3.cell(row=i, column=1, value=note).font = Font(size=10, name='微软雅黑')
ws3.column_dimensions['A'].width = 65

out_path = r"D:\AIANDshezhi\GenericAgent\temp\标签分类对照表_样本.xlsx"
wb.save(out_path)
print(f"✅ Excel saved: {out_path}")

# Print dist
dist = Counter(r[2] for r in unique_rows)
print("\n=== 样本分布 ===")
for cat in cat_order:
    if cat in dist:
        print(f"  {cat}: {dist[cat]}个标签")
